"""Import "GWG Pricing Master Formula.xlsx" into the v2 pricing config.

Reads base rates from the Settings tab rather than the Screen Print / Embroidery
/ DTF tabs, because those tabs are display copies with the multipliers already
baked in (the Embroidery tab even multiplies the digitizing fee, which must stay
a flat pass-through).

The garment markup sheet is 150 cost rows x 1000 quantity columns, but every
cell is linear between the quantity anchors, so only the anchors are exported.

Writes:
  docs/pricing/gwg-pricing-master.json                    (reference / diffable)
  packages/pricing/src/v2/generated/pricing-master.ts     (engine default config)

Usage:
  python3 scripts/import-pricing-master.py [path/to/GWG Pricing Master Formula.xlsx]
"""

import json
import math
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
JSON_OUT = REPO_ROOT / "docs" / "pricing" / "gwg-pricing-master.json"
TS_OUT = REPO_ROOT / "packages" / "pricing" / "src" / "v2" / "generated" / "pricing-master.ts"

QTY_ANCHORS_SCREEN = [1, 6, 12, 24, 48, 72, 144, 288, 500, 1000]
QTY_ANCHORS_RUN = [1, 6, 12, 24, 48, 72, 144, 288]

# Client-directed values that differ from the workbook's current cells. The
# workbook still charges $30 for new artwork (same as repeat) and applies rush
# to shipping; both were corrected in the Aug 2026 pricing review.
SETUP_NEW_PER_COLOUR_MINOR = 3500
DARK_GARMENT_PREMIUM_PERCENT = 0.10


def cents(value):
    return int(round(float(value) * 100))


def load(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    s = wb["Settings"]
    gm = wb["Garment Markup"]

    def markup(cost_dollars, qty):
        row = 3 + min(150, max(1, math.ceil(cost_dollars)))
        col = 1 + min(1000, max(1, round(qty)))
        return gm.cell(row, col).value

    # Settings!E12:N19 — screen print base rates, 8 colour rows x 10 qty bands.
    # Column D holds the colour count label, so rates start at column E (5).
    screen_rates = {
        str(colour): [cents(s.cell(11 + colour, 5 + i).value) for i in range(10)]
        for colour in range(1, 9)
    }
    # Settings!F24:H31 — embroidery base / extra per 1,000 / digitizing.
    embroidery_base = [cents(s.cell(23 + i, 6).value) for i in range(1, 9)]
    embroidery_extra = [cents(s.cell(23 + i, 7).value) for i in range(1, 9)]
    digitizing = cents(s.cell(24, 8).value)
    # Settings!F36:I43 — DTF rates by size.
    dtf = {
        key: [cents(s.cell(35 + i, col).value) for i in range(1, 9)]
        for key, col in (("small", 6), ("medium", 7), ("large", 8), ("oversize", 9))
    }

    return {
        "source": Path(xlsx_path).name,
        "settings": {
            "currency": "CAD",
            "minimumOrderQty": int(s["B4"].value),
            "rushFeePercent": float(s["B9"].value),
            "rushAppliesTo": "productionExcludingShipping",
            "packingFeePerGarmentMinor": cents(s["B11"].value),
            "shippingMarkupPercent": float(s["B12"].value),
            "quoteValidityDays": 30,
            "artworkMinimumFeeMinor": cents(s["B7"].value),
            "designHourlyRateMinor": cents(s["B8"].value),
            "darkGarmentRule": "everythingExceptWhite",
            "marginWarningThreshold": 0.35,
        },
        "garment": {
            "multiplier": float(s["E8"].value),
            "roundCostUpToWholeDollar": True,
            "costCapForMarkupMinor": 15000,
            "markupGrid": {
                "costAnchors": list(range(1, 151)),
                "qtyAnchors": QTY_ANCHORS_SCREEN,
                "grid": [
                    [markup(cost, qty) for qty in QTY_ANCHORS_SCREEN]
                    for cost in range(1, 151)
                ],
            },
            "mapPolicy": "warnOnly",
            "catalogDisplayQty": 24,
        },
        "methods": [
            {
                "key": "screenPrint",
                "label": "Screen print",
                "description": "Ink pushed through a screen. Best value at volume.",
                "enabled": True,
                "sortOrder": 1,
                "multiplier": float(s["E4"].value),
                "rateModel": {
                    "kind": "matrixByColour",
                    "qtyAnchors": QTY_ANCHORS_SCREEN,
                    "minColours": 1,
                    "maxColours": 8,
                    "ratesByColour": screen_rates,
                },
                "setup": {
                    "label": "Screen setup",
                    "description": "One screen burned per colour, per location.",
                    "newFeeMinor": SETUP_NEW_PER_COLOUR_MINOR,
                    "repeatFeeMinor": cents(s["B6"].value),
                    "per": "colour",
                    "shareAcrossGarments": True,
                    "multiplierApplies": False,
                    "repeatRequiresVerification": True,
                },
                "minimumChargePerLocationMinor": 0,
                "surcharges": [
                    {
                        "key": "darkGarment",
                        "label": "Dark garment",
                        "description": "Extra underbase ink on anything but white.",
                        "kind": "percent",
                        "value": DARK_GARMENT_PREMIUM_PERCENT,
                        "appliesWhen": "garmentIsDark",
                        "enabled": True,
                    },
                    {
                        "key": "oversized",
                        "label": "Oversized print",
                        "description": "Prints larger than a standard platen.",
                        "kind": "flatPerPiece",
                        "value": cents(s["B10"].value),
                        "appliesWhen": "locationFlagged",
                        "enabled": True,
                    },
                ],
                "costModel": {"runCostRatio": 0.45, "setupCostRatio": 0.2},
            },
            {
                "key": "embroidery",
                "label": "Embroidery",
                "description": "Stitched thread. Premium look, priced by stitch count.",
                "enabled": True,
                "sortOrder": 2,
                "multiplier": float(s["E6"].value),
                "rateModel": {
                    "kind": "baseWithVariable",
                    "qtyAnchors": QTY_ANCHORS_RUN,
                    "baseMinor": embroidery_base,
                    "extraPerUnitMinor": embroidery_extra,
                    "variable": {
                        "key": "stitchCount",
                        "label": "Stitches",
                        "unitSize": 1000,
                        "includedUnits": 5000,
                        "roundUpPartialUnits": False,
                    },
                },
                "setup": {
                    "label": "Digitizing",
                    "description": "Converting artwork to a stitch file. Once per logo.",
                    "newFeeMinor": digitizing,
                    "repeatFeeMinor": 0,
                    "per": "design",
                    "shareAcrossGarments": True,
                    "multiplierApplies": False,
                    "repeatRequiresVerification": True,
                },
                "minimumChargePerLocationMinor": 0,
                "surcharges": [
                    {
                        "key": "oversized",
                        "label": "Oversized design",
                        "description": "Designs beyond a standard hoop.",
                        "kind": "flatPerPiece",
                        "value": cents(s["B10"].value),
                        "appliesWhen": "locationFlagged",
                        "enabled": True,
                    }
                ],
                "costModel": {"runCostRatio": 0.4, "setupCostRatio": 0.2},
            },
            {
                "key": "dtf",
                "label": "DTF transfer",
                "description": "Direct-to-film. Full colour, no setup, priced by size.",
                "enabled": True,
                "sortOrder": 3,
                "multiplier": float(s["E7"].value),
                "rateModel": {
                    "kind": "matrixByOption",
                    "qtyAnchors": QTY_ANCHORS_RUN,
                    "options": [
                        {"key": "small", "label": "Small (up to 4\")"},
                        {"key": "medium", "label": "Medium (up to 8\")"},
                        {"key": "large", "label": "Large (up to 12\")"},
                        {"key": "oversize", "label": "Oversize (12\"+)"},
                    ],
                    "ratesByOption": dtf,
                },
                "setup": {
                    "label": "DTF setup",
                    "description": "No setup fee on DTF.",
                    "newFeeMinor": 0,
                    "repeatFeeMinor": 0,
                    "per": "design",
                    "shareAcrossGarments": True,
                    "multiplierApplies": False,
                    "repeatRequiresVerification": False,
                },
                "minimumChargePerLocationMinor": cents(s["B13"].value),
                "surcharges": [
                    {
                        "key": "oversized",
                        "label": "Oversized transfer",
                        "description": "Applies on top of the size rate when flagged.",
                        "kind": "flatPerPiece",
                        "value": cents(s["B10"].value),
                        "appliesWhen": "locationFlagged",
                        "enabled": True,
                    }
                ],
                "costModel": {"runCostRatio": 0.4, "setupCostRatio": 0.0},
            },
        ],
    }


def to_typescript(data):
    config = {
        "schemaVersion": 2,
        "version": 1,
        "status": "published",
        "effectiveFrom": "2026-08-01",
        "notes": f"Imported from {data['source']} on the Aug 2026 pricing review.",
        "settings": data["settings"],
        "garment": data["garment"],
        "methods": data["methods"],
    }
    body = json.dumps(config, indent=2)
    return (
        "// Generated by scripts/import-pricing-master.py — do not edit by hand.\n"
        "// Re-run the importer to pick up a new version of the estimator workbook.\n"
        'import type { PricingConfigV2 } from "@gwg/contracts";\n\n'
        "export const PRICING_MASTER_V2: PricingConfigV2 = "
        + body
        + " as PricingConfigV2;\n"
    )


def main(xlsx_path):
    data = load(xlsx_path)
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(data, indent=2) + "\n")
    TS_OUT.parent.mkdir(parents=True, exist_ok=True)
    TS_OUT.write_text(to_typescript(data))
    print(f"wrote {JSON_OUT.relative_to(REPO_ROOT)}")
    print(f"wrote {TS_OUT.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    default = Path.home() / "Downloads" / "GWG Pricing Master Formula.xlsx"
    main(sys.argv[1] if len(sys.argv) > 1 else str(default))
